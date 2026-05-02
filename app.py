from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from html import escape
from io import BytesIO
import json
from pathlib import Path

import streamlit as st

from judgelab.importer import ImportOptions, import_excel_files, preflight_excel_files
from judgelab.llm import LlmJudgeConfig, call_chat_json
from judgelab.models.macbert import EvaluateConfig, PredictConfig, TrainConfig, evaluate_saved_model, predict_records, train_classifier
from judgelab.workspace import Dataset, DatasetStats, WorkspaceStore, quote_ident
from judgelab.workflow import (
    append_records_table,
    clear_downstream_assets,
    count_filtered_records,
    create_filtered_sample,
    create_random_sample,
    create_splits,
    create_stratified_sample,
    fetch_table_records,
    get_asset_stats,
    label_distribution,
    preview_table,
    replace_records_table,
)


PROJECT_ROOT = Path(__file__).resolve().parent
WORKSPACE_DIR = PROJECT_ROOT / "judgelab_workspace"
STORE = WorkspaceStore(WORKSPACE_DIR)

FLOW_STEPS = [
    ("raw", "DATABASE", "原始数据"),
    ("sampled", "FILTER", "样本池"),
    ("labeled", "TAGS", "标注集"),
    ("splits", "SPLIT", "训练集"),
    ("model", "BRAIN", "模型"),
    ("predictions", "RESULT", "结果"),
]

VALID_STEPS = {key for key, _, _ in FLOW_STEPS}
FIELD_TYPE_LABELS = {
    "布尔": "boolean",
    "文本": "string",
    "数字": "number",
    "整数": "integer",
    "枚举": "enum",
}
FIELD_TYPE_VALUES = {value: label for label, value in FIELD_TYPE_LABELS.items()}
DEFAULT_LLM_FIELDS = [
    {
        "name": "is_target",
        "type_label": "布尔",
        "requirement": "判断文本是否属于目标类别",
        "example": "true",
        "required": True,
        "enum_options": "",
    },
    {
        "name": "core_reason",
        "type_label": "文本",
        "requirement": "用一句话说明判定理由",
        "example": "涉及低空物流配送场景",
        "required": True,
        "enum_options": "",
    },
    {
        "name": "confidence",
        "type_label": "数字",
        "requirement": "输出 0 到 1 之间的置信度",
        "example": "0.86",
        "required": True,
        "enum_options": "",
    },
]
DEFAULT_UI_CONFIG = {
    "sampling": {
        "method": "随机抽样",
        "sample_size": 1000,
        "seed": 42,
        "strata_column": "",
        "filter_columns": [],
        "filter_keywords": "",
        "keyword_mode": "任意关键词命中",
        "field_mode": "任意字段命中",
    },
    "llm": {
        "text_column": "",
        "api_key": "",
        "api_url": "https://open.bigmodel.cn/api/paas/v4/chat/completions",
        "model": "glm-5.1",
        "concurrency": 5,
        "limit": 100,
        "task_prompt": "请判断文本是否属于目标类别，并给出核心理由。",
        "fields": DEFAULT_LLM_FIELDS,
        "quality_label_column": "",
    },
    "split": {"source_table": "labeled_records", "label_column": "", "train_ratio": 0.8, "val_ratio": 0.1},
    "training": {
        "engine": "通用文本判定模型",
        "base_model": "hfl/chinese-macbert-base",
        "text_column": "",
        "label_column": "",
        "max_length": 256,
        "batch_size": 8,
        "epochs": 3,
        "learning_rate": 2e-5,
        "warmup_ratio": 0.1,
        "weight_decay": 0.01,
    },
}


def main() -> None:
    st.set_page_config(page_title="BigSample JudgeLab", layout="wide")
    inject_styles()
    sync_url_state()
    st.title("BigSample JudgeLab")
    st.caption("大样本智能判析实验室：多文件导入、偏向抽样、LLM 标注、质检、模型训练、全量预测。")

    selected_dataset = render_sidebar()
    if selected_dataset is None:
        render_empty_home()
        return

    render_dataset_workspace(selected_dataset)


def sync_url_state() -> None:
    dataset_id = query_param_value("dataset_id")
    if dataset_id:
        st.session_state["selected_dataset_id"] = dataset_id


def query_param_value(name: str) -> str | None:
    value = st.query_params.get(name)
    if isinstance(value, list):
        return value[0] if value else None
    return value


def ui_config_path(dataset_id: str) -> Path:
    return STORE.dataset_dir(dataset_id) / "ui_config.json"


def clone_config(value):
    return json.loads(json.dumps(value, ensure_ascii=False))


def merge_config(defaults: dict, saved: dict | None) -> dict:
    config = clone_config(defaults)
    if not isinstance(saved, dict):
        return config
    for key, value in saved.items():
        if isinstance(value, dict) and isinstance(config.get(key), dict):
            config[key].update(value)
        else:
            config[key] = value
    return config


def load_ui_config(dataset_id: str) -> dict:
    cache_key = f"ui_config_{dataset_id}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    path = ui_config_path(dataset_id)
    saved = None
    if path.exists():
        try:
            saved = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            saved = None
    config = merge_config(DEFAULT_UI_CONFIG, saved)
    st.session_state[cache_key] = config
    return config


def save_ui_config(dataset_id: str, config: dict) -> None:
    st.session_state[f"ui_config_{dataset_id}"] = config
    path = ui_config_path(dataset_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def update_step_config(dataset_id: str, step: str, values: dict) -> None:
    config = load_ui_config(dataset_id)
    current = config.setdefault(step, {})
    changed = False
    for key, value in values.items():
        if current.get(key) != value:
            current[key] = value
            changed = True
    if changed:
        save_ui_config(dataset_id, config)


def option_index(options: list, value, default: int = 0) -> int:
    try:
        return options.index(value)
    except ValueError:
        return default


def parse_keywords(text: str) -> list[str]:
    keywords: list[str] = []
    seen = set()
    for raw in text.replace("，", "\n").replace(",", "\n").splitlines():
        keyword = raw.strip()
        if keyword and keyword not in seen:
            keywords.append(keyword)
            seen.add(keyword)
    return keywords


def keyword_mode_value(label: str) -> str:
    return "all" if label == "全部关键词命中" else "any"


def field_mode_value(label: str) -> str:
    return "same_field" if label == "同一字段同时命中" else "any"


def render_sidebar() -> Dataset | None:
    st.sidebar.header("数据集")
    datasets = STORE.list_datasets()
    dataset_options = {f"{item.name} · {item.row_count:,} 行": item.dataset_id for item in datasets}

    if dataset_options:
        labels = list(dataset_options.keys())
        selected_dataset_id = st.session_state.get("selected_dataset_id")
        default_index = 0
        if selected_dataset_id:
            for index, label in enumerate(labels):
                if dataset_options[label] == selected_dataset_id:
                    default_index = index
                    break
        selected_label = st.sidebar.selectbox("选择数据集", labels, index=default_index)
        st.session_state["selected_dataset_id"] = dataset_options[selected_label]

    with st.sidebar.expander("新建数据集", expanded=not bool(datasets)):
        with st.form("create_dataset"):
            name = st.text_input("数据集名称", placeholder="例如：低空经济专利 2020-2025")
            description = st.text_area("说明", placeholder="可选：数据来源、分析目标、口径说明")
            submitted = st.form_submit_button("创建数据集", type="primary")
            if submitted:
                if not name.strip():
                    st.warning("请填写数据集名称。")
                else:
                    dataset = STORE.create_dataset(name.strip(), description.strip())
                    st.session_state["selected_dataset_id"] = dataset.dataset_id
                    st.session_state["active_step"] = "raw"
                    st.rerun()

    dataset_id = st.session_state.get("selected_dataset_id")
    if not dataset_id:
        return None
    try:
        dataset = STORE.get_dataset(dataset_id)
        render_dataset_actions(dataset)
        return dataset
    except KeyError:
        st.session_state.pop("selected_dataset_id", None)
        return None


def render_dataset_actions(dataset: Dataset) -> None:
    st.sidebar.divider()
    st.sidebar.subheader("数据集管理")
    with st.sidebar.expander("导入历史"):
        render_history_panel(dataset, compact=True)
    with st.sidebar.expander("清空/重置"):
        render_clear_panel(dataset, compact=True)


def render_empty_home() -> None:
    st.info("先在左侧创建一个数据集。一个数据集就是一条可追溯的大样本工作流。")
    st.markdown(
        """
        **推荐工作方式**

        1. 创建数据集
        2. 分批导入多个 Excel
        3. 系统检查表头一致性
        4. 数据进入 DuckDB，后续步骤不再反复读取 Excel
        5. 每次抽样、标注、训练、预测都记录为工作流产物
        """
    )


def render_dataset_workspace(dataset: Dataset) -> None:
    stats = STORE.get_dataset_stats(dataset.dataset_id)
    asset_stats = get_asset_stats(STORE, dataset.dataset_id)

    render_dataset_header(dataset, stats)

    active_step = st.session_state.get("active_step", "raw")
    if active_step == "quality":
        active_step = "labeled"
        st.session_state["active_step"] = active_step
    if active_step not in VALID_STEPS:
        active_step = "raw"
        st.session_state["active_step"] = active_step

    render_flow_nav(dataset, stats, asset_stats)
    st.divider()

    if active_step == "raw":
        render_import_panel(dataset, stats)
    elif active_step == "sampled":
        render_sampling_panel(dataset, stats)
    elif active_step == "labeled":
        render_llm_panel(dataset, stats, asset_stats)
    elif active_step == "splits":
        render_split_panel(dataset, stats, asset_stats)
    elif active_step == "model":
        render_training_panel(dataset, asset_stats)
    elif active_step == "predictions":
        render_prediction_panel(dataset, asset_stats)

    render_step_outputs(dataset, active_step)


def render_dataset_header(dataset: Dataset, stats: DatasetStats) -> None:
    st.markdown(
        f"""
        <section class="jl-dataset-header">
            <div>
                <h2>{escape(dataset.name)}</h2>
                <p>{escape(dataset.description or "暂无说明")}</p>
            </div>
            <div class="jl-metrics">
                <div class="jl-metric">
                    <span>原始数据行数</span>
                    <strong>{stats.row_count:,}</strong>
                </div>
                <div class="jl-metric">
                    <span>字段数</span>
                    <strong>{stats.column_count}</strong>
                </div>
            </div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_flow_nav(dataset: Dataset, stats: DatasetStats, asset_stats: dict[str, int]) -> None:
    raw_state = "done" if stats.raw_data_ready else "todo"
    model_result_ready = latest_training_result(dataset.dataset_id) is not None
    node_payload = {
        "raw": (raw_state, f"{stats.row_count:,} 行" if stats.raw_data_ready else "等待导入"),
        "sampled": ("done" if asset_stats.get("sampled", 0) else ("ready" if stats.raw_data_ready else "locked"), f"{asset_stats.get('sampled', 0):,} 行" if asset_stats.get("sampled", 0) else "待抽样"),
        "labeled": ("done" if asset_stats.get("reviewed", 0) else ("ready" if asset_stats.get("sampled", 0) else "locked"), f"质检 {asset_stats.get('reviewed', 0):,} 行" if asset_stats.get("reviewed", 0) else (f"初标 {asset_stats.get('labeled', 0):,} 行" if asset_stats.get("labeled", 0) else "待初标")),
        "splits": ("done" if asset_stats.get("train", 0) else ("ready" if asset_stats.get("reviewed", 0) or asset_stats.get("labeled", 0) or asset_stats.get("sampled", 0) else "locked"), f"训练 {asset_stats.get('train', 0):,} 行" if asset_stats.get("train", 0) else "待划分"),
        "model": ("done" if model_result_ready else ("ready" if asset_stats.get("train", 0) else "locked"), "已训练" if model_result_ready else "待训练"),
        "predictions": ("done" if asset_stats.get("predictions", 0) else ("ready" if model_result_ready else "locked"), f"{asset_stats.get('predictions', 0):,} 行" if asset_stats.get("predictions", 0) else "待判定"),
    }
    active_step = st.session_state.get("active_step", "raw")
    with st.container(key="flow_nav"):
        column_widths = []
        for index in range(len(FLOW_STEPS)):
            column_widths.append(1)
            if index < len(FLOW_STEPS) - 1:
                column_widths.append(0.12)
        columns = st.columns(column_widths, gap="small")
        button_index = 0
        for index, (key, eyebrow, label) in enumerate(FLOW_STEPS):
            state, caption = node_payload[key]
            disabled = state == "locked"
            active = active_step == key
            button_label = f"{eyebrow}\n\n{label}\n\n{caption}"
            with columns[button_index]:
                visual_state = "active" if active else ("ready" if state == "done" else "notready")
                with st.container(key=f"flow_node_{key}_{visual_state}"):
                    if st.button(
                        button_label,
                        key=f"flow_step_{dataset.dataset_id}_{key}",
                        disabled=disabled,
                        type="secondary",
                        width="stretch",
                    ):
                        st.session_state["active_step"] = key
                        st.rerun()
            if index < len(FLOW_STEPS) - 1:
                with columns[button_index + 1]:
                    st.markdown('<div class="jl-flow-arrow">→</div>', unsafe_allow_html=True)
            button_index += 2


def render_import_panel(dataset: Dataset, stats: DatasetStats) -> None:
    st.markdown("#### 多 Excel 分批导入")
    st.caption("第一次导入会建立标准表头；后续导入会按这个表头做一致性检查。字段一致但顺序不同可以自动重排。")
    st.info(f"当前数据集已有原始数据：{stats.row_count:,} 行。补充导入会追加到当前数据集，不会覆盖已导入数据。")
    render_import_success_list(dataset)

    upload_key = f"import_upload_{dataset.dataset_id}_{st.session_state.get('import_upload_version', 0)}"
    uploaded_files = st.file_uploader("选择一个或多个 Excel 文件", type=["xlsx", "xlsm", "xltx", "xltm"], accept_multiple_files=True, key=upload_key)
    if not uploaded_files:
        if stats.schema:
            st.info("当前标准表头：" + "，".join(stats.schema[:20]) + (" ..." if len(stats.schema) > 20 else ""))
        return

    staging_dir = STORE.dataset_dir(dataset.dataset_id) / "files" / "staging"
    staging_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    for uploaded in uploaded_files:
        target = staging_dir / Path(uploaded.name).name
        target.write_bytes(uploaded.getbuffer())
        paths.append(target)

    report = preflight_excel_files(paths, stats.schema or None)
    render_preflight_report(report)

    allow_extra = st.checkbox("允许导入多出的字段，并扩展数据集表头", value=False)
    allow_missing = st.checkbox("允许缺失字段导入，缺失列填空", value=False)
    has_extra = any(item.status == "extra_columns" for item in report.files)
    has_missing = any(item.status == "missing_columns" for item in report.files)
    has_empty_header = any(item.status == "empty_header" for item in report.files)
    import_disabled = not report.files or has_empty_header or (has_extra and not allow_extra) or (has_missing and not allow_missing)
    progress = st.progress(0, text="等待导入确认 0%")

    if st.button("导入到当前数据集", type="primary", disabled=import_disabled):
        progress.progress(1, text="正在准备导入 1%")

        def on_progress(done: int, total: int, message: str) -> None:
            percent = min(99, max(1, int(done / max(1, total) * 100)))
            progress.progress(percent, text=f"{message} {percent}%")

        try:
            summary = import_excel_files(
                STORE,
                dataset.dataset_id,
                paths,
                ImportOptions(allow_extra_columns=allow_extra, allow_missing_columns=allow_missing),
                progress_callback=on_progress,
            )
        except Exception as exc:
            progress.empty()
            st.error(str(exc))
            return
        progress.progress(100, text="导入完成 100%")
        clear_downstream_assets(STORE, dataset.dataset_id, "raw")
        st.success(f"导入完成：{summary.imported_files} 个文件，{summary.imported_rows:,} 行。")
        st.session_state["last_import_result"] = {"files": summary.imported_files, "rows": summary.imported_rows}
        st.session_state["import_upload_version"] = st.session_state.get("import_upload_version", 0) + 1
        for path in paths:
            path.unlink(missing_ok=True)
        st.rerun()


def render_import_success_list(dataset: Dataset) -> None:
    result = st.session_state.get("last_import_result")
    if result:
        st.success(f"最近一次导入成功：{result['files']} 个文件，{result['rows']:,} 行。现在可以继续补充导入。")
    batches = STORE.list_import_batches(dataset.dataset_id, limit=8)
    if not batches:
        return
    rows = [
        {"文件": item.file_name, "状态": "导入成功" if item.status == "imported" else item.status, "行数": item.row_count, "时间": item.created_at}
        for item in batches
    ]
    st.dataframe(rows, width="stretch")


def render_preflight_report(report) -> None:
    status_label = {
        "matched": "通过",
        "reordered": "顺序不同",
        "missing_columns": "缺少字段",
        "extra_columns": "多出字段",
        "empty_header": "表头为空",
    }
    rows = [
        {
            "文件": item.file_name,
            "状态": status_label.get(item.status, item.status),
            "行数": item.row_count,
            "缺少字段": "，".join(item.missing_columns),
            "多出字段": "，".join(item.extra_columns),
            "说明": item.message,
        }
        for item in report.files
    ]
    c1, c2, c3 = st.columns(3)
    c1.metric("文件数", len(report.files))
    c2.metric("预计导入行数", f"{report.total_rows:,}")
    c3.metric("严格模式", "可导入" if report.can_import_strict else "需处理")
    st.dataframe(rows, width="stretch")


def render_pagination(page_key: str, max_page: int, dataset_id: str | None = None) -> None:
    st.markdown("##### 数据分页")
    prev_col, slider_col, next_col, jump_col = st.columns([1, 5, 1, 2], vertical_alignment="bottom")
    with prev_col:
        if st.button("‹ 上一页", disabled=st.session_state[page_key] <= 1, key=f"{page_key}_prev", width="stretch"):
            st.session_state[page_key] -= 1
            st.rerun()
    with slider_col:
        if max_page > 1:
            page = st.slider(
                "页码",
                min_value=1,
                max_value=max_page,
                value=int(st.session_state[page_key]),
                step=1,
                key=f"{page_key}_slider",
                label_visibility="collapsed",
            )
            if page != st.session_state[page_key]:
                st.session_state[page_key] = int(page)
                st.rerun()
        else:
            st.caption("当前只有 1 页")
    with next_col:
        if st.button("下一页 ›", disabled=st.session_state[page_key] >= max_page, key=f"{page_key}_next", width="stretch"):
            st.session_state[page_key] += 1
            st.rerun()
    with jump_col:
        jump_key = f"{page_key}_jump_value"
        jump_value = st.text_input("跳转", value=str(st.session_state[page_key]), key=jump_key, label_visibility="collapsed")
        if st.button("跳转到页", key=f"{page_key}_jump", width="stretch", disabled=max_page <= 1):
            try:
                target = int(jump_value)
            except ValueError:
                target = st.session_state[page_key]
            st.session_state[page_key] = min(max(1, target), max_page)
            st.rerun()


def render_sampling_panel(dataset: Dataset, stats: DatasetStats) -> None:
    st.markdown("#### 智能抽样")
    if not stats.raw_data_ready:
        st.info("请先导入原始数据。")
        return
    config = load_ui_config(dataset.dataset_id).get("sampling", {})
    method_options = ["随机抽样", "分层抽样", "筛选后抽样"]
    default_sample_size = min(max(1, int(config.get("sample_size", 1000) or 1000)), max(1, stats.row_count))
    col1, col2, col3 = st.columns(3)
    with col1:
        method = st.selectbox("抽样方式", method_options, index=option_index(method_options, config.get("method"), 0), key=f"sampling_method_{dataset.dataset_id}")
    with col2:
        sample_size = st.number_input("样本量", min_value=1, max_value=max(1, stats.row_count), value=default_sample_size, step=100, key=f"sampling_size_{dataset.dataset_id}")
    with col3:
        seed = st.number_input("随机种子", min_value=0, value=int(config.get("seed", 42) or 42), step=1, key=f"sampling_seed_{dataset.dataset_id}")
    strata_column = None
    if method == "分层抽样":
        strata_column = st.selectbox("分层字段", stats.schema, index=option_index(stats.schema, config.get("strata_column"), 0), key=f"sampling_strata_{dataset.dataset_id}")
    filter_columns: list[str] = []
    filter_keywords = str(config.get("filter_keywords", "") or "")
    keyword_mode = str(config.get("keyword_mode", "任意关键词命中") or "任意关键词命中")
    field_mode = str(config.get("field_mode", "任意字段命中") or "任意字段命中")
    matched_count: int | None = None
    if method == "筛选后抽样":
        st.markdown("##### 筛选条件")
        text_like_defaults = [column for column in stats.schema if any(token in column for token in ["标题", "摘要", "文本", "内容", "权利", "说明", "关键词"])]
        saved_columns = [column for column in config.get("filter_columns", []) if column in stats.schema] if isinstance(config.get("filter_columns"), list) else []
        default_columns = saved_columns or text_like_defaults[: min(3, len(text_like_defaults))] or stats.schema[:1]
        filter_columns = st.multiselect("筛选字段", stats.schema, default=default_columns, key=f"sampling_filter_columns_{dataset.dataset_id}")
        filter_keywords = st.text_area(
            "关键词",
            value=filter_keywords,
            height=120,
            placeholder="每行一个关键词，例如：\n无人机\neVTOL\n低空物流",
            key=f"sampling_filter_keywords_{dataset.dataset_id}",
        )
        c1, c2 = st.columns(2)
        with c1:
            keyword_options = ["任意关键词命中", "全部关键词命中"]
            keyword_mode = st.selectbox("关键词逻辑", keyword_options, index=option_index(keyword_options, keyword_mode, 0), key=f"sampling_keyword_mode_{dataset.dataset_id}")
        with c2:
            field_options = ["任意字段命中", "同一字段同时命中"]
            field_mode = st.selectbox("字段逻辑", field_options, index=option_index(field_options, field_mode, 0), key=f"sampling_field_mode_{dataset.dataset_id}")
        keywords = parse_keywords(filter_keywords)
        if filter_columns and keywords:
            try:
                matched_count = count_filtered_records(
                    STORE,
                    dataset.dataset_id,
                    filter_columns,
                    keywords,
                    keyword_mode=keyword_mode_value(keyword_mode),
                    field_mode=field_mode_value(field_mode),
                )
                st.metric("筛选命中行数", f"{matched_count:,}")
            except Exception as exc:
                st.warning(str(exc))
        else:
            st.info("选择字段并输入关键词后，系统会先预估命中行数，再从命中结果中随机抽样。")
    update_step_config(
        dataset.dataset_id,
        "sampling",
        {
            "method": method,
            "sample_size": int(sample_size),
            "seed": int(seed),
            "strata_column": strata_column or config.get("strata_column", ""),
            "filter_columns": filter_columns or config.get("filter_columns", []),
            "filter_keywords": filter_keywords,
            "keyword_mode": keyword_mode,
            "field_mode": field_mode,
        },
    )
    progress = st.progress(0, text="等待抽样 0%")
    if st.button("生成样本池", type="primary"):
        progress.progress(30, text="正在从 DuckDB 抽样 30%")
        try:
            if method == "分层抽样":
                result = create_stratified_sample(STORE, dataset.dataset_id, strata_column, int(sample_size), int(seed))
            elif method == "筛选后抽样":
                keywords = parse_keywords(filter_keywords)
                result = create_filtered_sample(
                    STORE,
                    dataset.dataset_id,
                    filter_columns,
                    keywords,
                    int(sample_size),
                    int(seed),
                    keyword_mode=keyword_mode_value(keyword_mode),
                    field_mode=field_mode_value(field_mode),
                )
            else:
                result = create_random_sample(STORE, dataset.dataset_id, int(sample_size), int(seed))
        except Exception as exc:
            progress.empty()
            st.error(str(exc))
            return
        progress.progress(100, text="样本池生成完成 100%")
        if matched_count is not None:
            st.success(f"样本池已生成：{result.row_count:,} 行，来自 {matched_count:,} 条筛选命中数据。")
        else:
            st.success(f"样本池已生成：{result.row_count:,} 行。")
        st.dataframe(preview_table(STORE, dataset.dataset_id, "sampled_records", limit=50), width="stretch")


def render_split_panel(dataset: Dataset, stats: DatasetStats, asset_stats: dict[str, int]) -> None:
    st.markdown("#### 数据集划分")
    if not asset_stats.get("sampled", 0):
        st.info("请先生成样本池。后续 LLM 初标/人工复核完成后，可以选择标注列进行训练集划分。")
        return
    config = load_ui_config(dataset.dataset_id).get("split", {})
    source_options = ["reviewed_records", "labeled_records", "sampled_records"]
    if not asset_stats.get("reviewed", 0):
        source_options.remove("reviewed_records")
    if not asset_stats.get("labeled", 0):
        source_options.remove("labeled_records")
    source_table = st.selectbox("划分来源", source_options, index=option_index(source_options, config.get("source_table"), 0), key=f"split_source_{dataset.dataset_id}")
    table_columns = columns_for_table(dataset.dataset_id, source_table) or stats.schema
    label_column = st.selectbox("标签列", table_columns, index=option_index(table_columns, config.get("label_column"), 0), key=f"split_label_{dataset.dataset_id}")
    c1, c2, c3 = st.columns(3)
    with c1:
        train_ratio = st.slider("训练集", 0.5, 0.9, float(config.get("train_ratio", 0.8) or 0.8), 0.05, key=f"split_train_{dataset.dataset_id}")
    with c2:
        val_ratio = st.slider("验证集", 0.05, 0.3, float(config.get("val_ratio", 0.1) or 0.1), 0.05, key=f"split_val_{dataset.dataset_id}")
    with c3:
        test_ratio = round(1.0 - train_ratio - val_ratio, 2)
        st.metric("测试集", f"{test_ratio:.2f}")
    update_step_config(
        dataset.dataset_id,
        "split",
        {"source_table": source_table, "label_column": label_column, "train_ratio": float(train_ratio), "val_ratio": float(val_ratio)},
    )
    if test_ratio <= 0:
        st.error("训练集 + 验证集比例必须小于 1。")
        return
    progress = st.progress(0, text="等待划分 0%")
    if st.button("生成训练/验证/测试集", type="primary"):
        progress.progress(40, text="正在分层划分 40%")
        try:
            result = create_splits(STORE, dataset.dataset_id, source_table, label_column, train_ratio, val_ratio, test_ratio)
        except Exception as exc:
            progress.empty()
            st.error(str(exc))
            return
        progress.progress(100, text="数据集划分完成 100%")
        st.success(f"划分完成：训练 {result.train_count:,}，验证 {result.val_count:,}，测试 {result.test_count:,}。")


def render_llm_panel(dataset: Dataset, stats: DatasetStats, asset_stats: dict[str, int]) -> None:
    st.markdown("#### LLM 初标")
    if not asset_stats.get("sampled", 0):
        st.info("请先生成样本池。LLM 初标默认只处理样本池，不建议直接处理全量原始数据。")
        return
    config = load_ui_config(dataset.dataset_id).get("llm", {})
    text_column = st.selectbox(
        "文本列",
        stats.schema,
        index=option_index(stats.schema, config.get("text_column"), 0),
        key=f"llm_text_column_{dataset.dataset_id}",
    )
    api_key = st.text_input("API Key", type="password", value=str(config.get("api_key", "")), key=f"llm_api_key_{dataset.dataset_id}")
    api_url = st.text_input("API URL", value=str(config.get("api_url", DEFAULT_UI_CONFIG["llm"]["api_url"])), key=f"llm_api_url_{dataset.dataset_id}")
    model = st.text_input("模型", value=str(config.get("model", DEFAULT_UI_CONFIG["llm"]["model"])), key=f"llm_model_{dataset.dataset_id}")
    concurrency = st.number_input("并发数", min_value=1, max_value=50, value=int(config.get("concurrency", 5) or 5), step=1, key=f"llm_concurrency_{dataset.dataset_id}")
    max_limit = max(1, asset_stats.get("sampled", 1))
    default_limit = min(max_limit, max(1, int(config.get("limit", 100) or 100)))
    limit = st.number_input("本次最多处理样本数", min_value=1, max_value=max_limit, value=default_limit, step=10, key=f"llm_limit_{dataset.dataset_id}")
    task_prompt = st.text_area(
        "判定任务说明",
        value=str(config.get("task_prompt", DEFAULT_UI_CONFIG["llm"]["task_prompt"])),
        height=120,
        key=f"llm_task_prompt_{dataset.dataset_id}",
    )
    st.markdown("##### 结构化输出字段")
    st.caption("每一行就是大模型需要输出的一个 JSON 一级字段。字段名只能用英文、数字和下划线；不支持二级结构。")
    edited_fields = st.data_editor(
        field_rows_for_editor(config.get("fields")),
        hide_index=True,
        num_rows="dynamic",
        width="stretch",
        key=f"llm_fields_{dataset.dataset_id}",
        column_config={
            "name": st.column_config.TextColumn("输出字段名", help="例如 is_target、core_reason、confidence"),
            "type_label": st.column_config.SelectboxColumn("类型", options=list(FIELD_TYPE_LABELS.keys()), required=True),
            "requirement": st.column_config.TextColumn("提取/判断要求", help="告诉模型这个字段应该如何判断或提取"),
            "example": st.column_config.TextColumn("示例值", help="用户自己输入一个标准输出示例"),
            "required": st.column_config.CheckboxColumn("必填"),
            "enum_options": st.column_config.TextColumn("枚举选项", help="类型选择枚举时填写，例如 相关,不相关,不确定"),
        },
    )
    field_specs = normalize_field_rows(edited_fields)
    update_step_config(
        dataset.dataset_id,
        "llm",
        {
            "text_column": text_column,
            "api_key": api_key,
            "api_url": api_url,
            "model": model,
            "concurrency": int(concurrency),
            "limit": int(limit),
            "task_prompt": task_prompt,
            "fields": field_rows_for_editor(field_specs),
            "quality_label_column": config.get("quality_label_column", ""),
        },
    )
    system_prompt = build_structured_prompt(task_prompt, field_specs)
    with st.expander("预览最终 Prompt"):
        st.code(system_prompt)
    with st.expander("预览 JSON 输出结构"):
        preview_json = {field["name"]: example_value(str(field["type"]), str(field.get("example", "") or "")) for field in field_specs}
        st.json(preview_json)

    progress = st.progress(0, text="等待 LLM 初标 0%")
    if st.button("开始 LLM 初标", type="primary"):
        if not field_specs:
            st.error("请至少配置一个结构化输出字段。")
            return
        records = fetch_table_records(STORE, dataset.dataset_id, "sampled_records", limit=int(limit))
        if not records:
            st.error("样本池为空。")
            return
        config = LlmJudgeConfig(api_key=api_key, api_url=api_url, model=model, system_prompt=system_prompt, text_column=text_column)
        labeled = []
        done = 0
        errors = 0
        with ThreadPoolExecutor(max_workers=int(concurrency)) as executor:
            future_map = {executor.submit(label_one_record, record, config, field_specs): record for record in records}
            for future in as_completed(future_map):
                labeled.append(future.result())
                done += 1
                if labeled[-1].get("LLM_ERROR"):
                    errors += 1
                percent = int(done / len(records) * 100)
                progress.progress(percent, text=f"LLM 初标中 {percent}% · 成功 {done - errors} · 失败 {errors}")
        replace_records_table(STORE, dataset.dataset_id, "labeled_records", labeled)
        clear_downstream_assets(STORE, dataset.dataset_id, "labeled")
        progress.progress(100, text="LLM 初标完成 100%")
        st.success(f"已写入标注集：{len(labeled):,} 行，失败 {errors:,} 行。")
        st.dataframe(preview_table(STORE, dataset.dataset_id, "labeled_records", limit=50), width="stretch")
        st.rerun()

    render_label_inspection_panel(dataset, asset_stats)


def render_label_inspection_panel(dataset: Dataset, asset_stats: dict[str, int]) -> None:
    st.divider()
    st.markdown("#### 标签质检")
    if not asset_stats.get("labeled", 0):
        st.info("完成 LLM 初标后，这里会展示标签分布、失败样本和高质量标注集。")
        return
    table_columns = columns_for_table(dataset.dataset_id, "labeled_records")
    config = load_ui_config(dataset.dataset_id).get("llm", {})
    label_column = st.selectbox(
        "质检标签列",
        table_columns,
        index=option_index(table_columns, config.get("quality_label_column"), 0),
        key=f"quality_label_{dataset.dataset_id}",
    )
    update_step_config(dataset.dataset_id, "llm", {"quality_label_column": label_column})
    distribution = label_distribution(STORE, dataset.dataset_id, "labeled_records", label_column)
    c1, c2, c3 = st.columns(3)
    c1.metric("标注集行数", f"{asset_stats.get('labeled', 0):,}")
    c2.metric("标签种类", len(distribution))
    c3.metric("未标注", distribution.get("未标注", 0))
    st.dataframe([{"标签": key, "数量": value} for key, value in distribution.items()], width="stretch")
    if "LLM_ERROR" in table_columns:
        st.caption("如需复核失败样本，可在标注集预览中查看 LLM_ERROR 列。")
    st.markdown("##### 高质量标注集")
    reviewed_count = asset_stats.get("reviewed", 0)
    if reviewed_count:
        st.success(f"高质量标注集已生成：{reviewed_count:,} 行。后续训练集划分会优先使用它。")
        st.dataframe(preview_table(STORE, dataset.dataset_id, "reviewed_records", limit=30), width="stretch")
        if st.button("重新生成高质量标注集", type="primary", key=f"refresh_reviewed_{dataset.dataset_id}"):
            records = fetch_table_records(STORE, dataset.dataset_id, "labeled_records", limit=int(asset_stats.get("labeled", 0)))
            replace_records_table(STORE, dataset.dataset_id, "reviewed_records", records)
            clear_downstream_assets(STORE, dataset.dataset_id, "reviewed")
            st.success(f"高质量标注集已更新：{len(records):,} 行。")
            st.rerun()
    else:
        st.info("质检通过后，可以把当前标注集确认为高质量标注集，供后续训练集划分使用。")
        if st.button("确认当前标注集通过质检", type="primary", key=f"approve_labeled_{dataset.dataset_id}"):
            records = fetch_table_records(STORE, dataset.dataset_id, "labeled_records", limit=int(asset_stats.get("labeled", 0)))
            replace_records_table(STORE, dataset.dataset_id, "reviewed_records", records)
            clear_downstream_assets(STORE, dataset.dataset_id, "reviewed")
            st.success(f"高质量标注集已生成：{len(records):,} 行。")
            st.rerun()


def render_training_panel(dataset: Dataset, asset_stats: dict[str, int]) -> None:
    st.markdown("#### 模型训练")
    if not asset_stats.get("train", 0):
        st.info("请先完成数据集划分。")
        return
    st.success(f"训练数据已就绪：训练 {asset_stats.get('train', 0):,} 行，验证 {asset_stats.get('val', 0):,} 行，测试 {asset_stats.get('test', 0):,} 行。")
    config = load_ui_config(dataset.dataset_id).get("training", {})
    train_columns = columns_for_table(dataset.dataset_id, "train_records")
    if not train_columns:
        st.error("训练集为空，无法训练。")
        return

    engine_options = ["通用文本判定模型"]
    engine = st.selectbox("训练模块", engine_options, index=option_index(engine_options, config.get("engine"), 0), key=f"train_engine_{dataset.dataset_id}")
    col1, col2 = st.columns(2)
    with col1:
        text_column = st.selectbox("文本字段", train_columns, index=option_index(train_columns, config.get("text_column"), 0), key=f"train_text_column_{dataset.dataset_id}")
    with col2:
        label_column = st.selectbox("标签字段", train_columns, index=option_index(train_columns, config.get("label_column"), 0), key=f"train_label_column_{dataset.dataset_id}")

    with st.expander("训练参数", expanded=True):
        st.caption("第一期默认使用中文文本判定模型，底层基础模型为 hfl/chinese-macbert-base；后续可以替换成其他训练引擎。")
        base_model = st.text_input("基础模型或本地模型目录", value=str(config.get("base_model", "hfl/chinese-macbert-base")), key=f"train_base_model_{dataset.dataset_id}")
        c1, c2, c3 = st.columns(3)
        with c1:
            epochs = st.number_input("训练轮数", min_value=1, max_value=20, value=int(config.get("epochs", 3) or 3), step=1, key=f"train_epochs_{dataset.dataset_id}")
            max_length = st.number_input("最大文本长度", min_value=64, max_value=1024, value=int(config.get("max_length", 256) or 256), step=32, key=f"train_max_length_{dataset.dataset_id}")
        with c2:
            batch_size = st.number_input("批大小", min_value=1, max_value=128, value=int(config.get("batch_size", 8) or 8), step=1, key=f"train_batch_size_{dataset.dataset_id}")
            learning_rate = st.number_input("学习率", min_value=0.000001, max_value=0.001, value=float(config.get("learning_rate", 2e-5) or 2e-5), step=0.000001, format="%.6f", key=f"train_lr_{dataset.dataset_id}")
        with c3:
            warmup_ratio = st.number_input("预热比例", min_value=0.0, max_value=0.5, value=float(config.get("warmup_ratio", 0.1) or 0.1), step=0.01, format="%.2f", key=f"train_warmup_{dataset.dataset_id}")
            weight_decay = st.number_input("权重衰减", min_value=0.0, max_value=0.5, value=float(config.get("weight_decay", 0.01) or 0.01), step=0.01, format="%.2f", key=f"train_weight_decay_{dataset.dataset_id}")

    update_step_config(
        dataset.dataset_id,
        "training",
        {
            "engine": engine,
            "base_model": base_model,
            "text_column": text_column,
            "label_column": label_column,
            "max_length": int(max_length),
            "batch_size": int(batch_size),
            "epochs": int(epochs),
            "learning_rate": float(learning_rate),
            "warmup_ratio": float(warmup_ratio),
            "weight_decay": float(weight_decay),
        },
    )

    latest_result = latest_training_result(dataset.dataset_id)
    if latest_result:
        with st.expander("最近一次训练结果", expanded=False):
            st.json(latest_result)

    if st.button("开始训练模型", type="primary", key=f"start_training_{dataset.dataset_id}"):
        output_dir = STORE.dataset_dir(dataset.dataset_id) / "models" / datetime.now().strftime("v%Y%m%d_%H%M%S")
        logs: list[str] = []
        log_box = st.empty()
        progress = st.progress(0, text="正在准备训练数据 0%")

        def append_log(message: str) -> None:
            logs.append(message)
            log_box.code("\n".join(logs[-30:]))

        try:
            train_records = fetch_table_records(STORE, dataset.dataset_id, "train_records")
            val_records = fetch_table_records(STORE, dataset.dataset_id, "val_records")
            test_records = fetch_table_records(STORE, dataset.dataset_id, "test_records")
            progress.progress(10, text="训练数据读取完成 10%")
            result = train_classifier(
                train_records,
                val_records,
                test_records,
                TrainConfig(
                    base_model=base_model,
                    output_dir=output_dir,
                    text_column=text_column,
                    label_column=label_column,
                    max_length=int(max_length),
                    batch_size=int(batch_size),
                    epochs=int(epochs),
                    learning_rate=float(learning_rate),
                    warmup_ratio=float(warmup_ratio),
                    weight_decay=float(weight_decay),
                    metadata={
                        "dataset_id": dataset.dataset_id,
                        "dataset_name": dataset.name,
                        "split": load_ui_config(dataset.dataset_id).get("split", {}),
                        "training": {
                            "engine": engine,
                            "base_model": base_model,
                            "text_column": text_column,
                            "label_column": label_column,
                            "max_length": int(max_length),
                            "batch_size": int(batch_size),
                            "epochs": int(epochs),
                            "learning_rate": float(learning_rate),
                            "warmup_ratio": float(warmup_ratio),
                            "weight_decay": float(weight_decay),
                        },
                    },
                ),
                log=append_log,
            )
        except Exception as exc:
            progress.empty()
            st.error(str(exc))
            return

        progress.progress(100, text="模型训练完成 100%")
        st.success(f"训练完成，模型已保存到：{output_dir}")
        st.json(result)


def latest_training_result(dataset_id: str) -> dict | None:
    training_dir = latest_training_dir(dataset_id)
    if training_dir is None:
        return None
    result_path = training_dir / "training_result.json"
    if not result_path.exists():
        return None
    try:
        return json.loads(result_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def latest_training_dir(dataset_id: str) -> Path | None:
    models_dir = STORE.dataset_dir(dataset_id) / "models"
    if not models_dir.exists():
        return None
    result_files = sorted(models_dir.glob("*/training_result.json"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not result_files:
        return None
    return result_files[0].parent


def list_training_dirs(dataset_id: str) -> list[Path]:
    models_dir = STORE.dataset_dir(dataset_id) / "models"
    if not models_dir.exists():
        return []
    return sorted(
        [path.parent for path in models_dir.glob("*/training_result.json")],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def load_training_result_from_dir(training_dir: Path) -> dict | None:
    result_path = training_dir / "training_result.json"
    if not result_path.exists():
        return None
    try:
        return json.loads(result_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def load_evaluation_details_from_dir(training_dir: Path) -> dict[str, list[dict]] | None:
    detail_path = training_dir / "evaluation_details.json"
    if not detail_path.exists():
        return None
    try:
        payload = json.loads(detail_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    return {
        "train": payload.get("train", []) if isinstance(payload, dict) else [],
        "validation": payload.get("validation", []) if isinstance(payload, dict) else [],
        "test": payload.get("test", []) if isinstance(payload, dict) else [],
    }


def latest_evaluation_details(dataset_id: str) -> dict[str, list[dict]] | None:
    training_dir = latest_training_dir(dataset_id)
    if training_dir is None:
        return None
    return load_evaluation_details_from_dir(training_dir)


def render_prediction_panel(dataset: Dataset, asset_stats: dict[str, int]) -> None:
    st.markdown("#### 结果验证")
    training_dirs = list_training_dirs(dataset.dataset_id)
    if not training_dirs:
        if not asset_stats.get("train", 0):
            st.info("请先完成数据集划分并训练模型。结果节点会展示模型对当前验证集/测试集的判定准确度。")
        else:
            st.info("请先完成模型训练。")
        return
    training_labels = [path.name for path in training_dirs]
    selected_label = st.selectbox("选择模型版本", training_labels, index=0, key=f"result_model_version_{dataset.dataset_id}")
    selected_dir = training_dirs[training_labels.index(selected_label)]
    result = load_training_result_from_dir(selected_dir) or {}
    details = load_evaluation_details_from_dir(selected_dir) or {"train": [], "validation": [], "test": []}
    metadata = result.get("metadata", {}) if isinstance(result, dict) else {}
    split_meta = metadata.get("split", {}) if isinstance(metadata, dict) else {}
    training_meta = metadata.get("training", {}) if isinstance(metadata, dict) else {}
    train_columns = columns_for_table(dataset.dataset_id, "train_records")
    text_column_options = train_columns or columns_for_table(dataset.dataset_id, "val_records") or columns_for_table(dataset.dataset_id, "test_records")
    default_text_column = str(training_meta.get("text_column", "") or (text_column_options[0] if text_column_options else ""))
    default_label_column = str(training_meta.get("label_column", "") or (text_column_options[0] if text_column_options else ""))

    config_col1, config_col2, config_col3 = st.columns(3)
    with config_col1:
        text_column = st.selectbox(
            "验证文本字段",
            text_column_options,
            index=option_index(text_column_options, default_text_column, 0) if text_column_options else 0,
            key=f"result_text_column_{dataset.dataset_id}_{selected_label}",
        ) if text_column_options else ""
    with config_col2:
        label_column = st.selectbox(
            "验证标签字段",
            text_column_options,
            index=option_index(text_column_options, default_label_column, 0) if text_column_options else 0,
            key=f"result_label_column_{dataset.dataset_id}_{selected_label}",
        ) if text_column_options else ""
    with config_col3:
        reevaluate_clicked = st.button("重新验证当前模型", key=f"reevaluate_model_{dataset.dataset_id}_{selected_label}", type="primary")
    st.caption("重新验证不会重新训练模型，只会使用当前数据集里的训练集、验证集、测试集重新计算结果。")

    if reevaluate_clicked:
        if not text_column or not label_column:
            st.error("请先选择文本字段和标签字段。")
            return
        train_records = fetch_table_records(STORE, dataset.dataset_id, "train_records")
        val_records = fetch_table_records(STORE, dataset.dataset_id, "val_records")
        test_records = fetch_table_records(STORE, dataset.dataset_id, "test_records")
        try:
            train_result = evaluate_saved_model(
                train_records,
                EvaluateConfig(
                    model_dir=selected_dir,
                    text_column=text_column,
                    label_column=label_column,
                    max_length=int(training_meta.get("max_length", 256) or 256),
                    batch_size=int(training_meta.get("batch_size", 16) or 16),
                ),
            )
            val_result = evaluate_saved_model(
                val_records,
                EvaluateConfig(
                    model_dir=selected_dir,
                    text_column=text_column,
                    label_column=label_column,
                    max_length=int(training_meta.get("max_length", 256) or 256),
                    batch_size=int(training_meta.get("batch_size", 16) or 16),
                ),
            )
            test_result = evaluate_saved_model(
                test_records,
                EvaluateConfig(
                    model_dir=selected_dir,
                    text_column=text_column,
                    label_column=label_column,
                    max_length=int(training_meta.get("max_length", 256) or 256),
                    batch_size=int(training_meta.get("batch_size", 16) or 16),
                ),
            )
        except Exception as exc:
            st.error(str(exc))
            return
        result["train_metrics"] = train_result["metrics"]
        result["best_val_metrics"] = val_result["metrics"]
        result["test_metrics"] = test_result["metrics"]
        result["train_size"] = len(train_result["details"])
        result["val_size"] = len(val_result["details"])
        result["test_size"] = len(test_result["details"])
        if not isinstance(result.get("metadata"), dict):
            result["metadata"] = {}
        result["metadata"]["training"] = {
            **training_meta,
            "text_column": text_column,
            "label_column": label_column,
        }
        (selected_dir / "training_result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        (selected_dir / "evaluation_details.json").write_text(
            json.dumps(
                {
                    "train": train_result["details"],
                    "validation": val_result["details"],
                    "test": test_result["details"],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        st.success("当前模型已重新验证，结果已更新。")
        st.rerun()

    st.info(
        "训练集参与模型学习，分数通常最高；验证集用于挑选最佳模型；测试集只用于最终验收。判断模型是否可用，优先看验证集和测试集。"
    )
    st.success(
        f"最近一次训练已完成：训练 {result.get('train_size', 0):,} 行，"
        f"验证 {result.get('val_size', 0):,} 行，测试 {result.get('test_size', 0):,} 行。"
    )
    render_model_assessment(result)

    render_metric_cards("训练集结果（看模型是否学会）", result.get("train_metrics", {}))
    render_metric_cards("验证集结果（用于选择模型）", result.get("best_val_metrics", {}))
    render_metric_cards("测试集结果（用于最终验收）", result.get("test_metrics", {}))

    with st.expander("本次验证配置", expanded=False):
        st.json(
            {
                "split": split_meta,
                "training": training_meta,
                "model_dir": result.get("model_dir", ""),
            }
        )

    render_evaluation_details("训练集逐条结果", details.get("train", []), key=f"train_details_{dataset.dataset_id}")
    render_evaluation_details("验证集逐条结果", details.get("validation", []), key=f"validation_details_{dataset.dataset_id}")
    render_evaluation_details("测试集逐条结果", details.get("test", []), key=f"test_details_{dataset.dataset_id}")
    render_full_prediction_panel(dataset, selected_dir, asset_stats, training_meta)


def render_metric_cards(title: str, metrics: dict) -> None:
    st.markdown(f"##### {title}")
    if not metrics:
        st.info("当前没有可展示的指标。")
        return
    metric_cols = st.columns(5)
    metric_cols[0].metric("Accuracy（准确率）", format_ratio(metrics.get("accuracy")))
    metric_cols[1].metric("Precision（精确率）", format_ratio(metrics.get("precision")))
    metric_cols[2].metric("Recall（召回率）", format_ratio(metrics.get("recall")))
    metric_cols[3].metric("F1（综合指标）", format_ratio(metrics.get("f1")))
    metric_cols[4].metric("样本数", f"{metric_support(metrics):,}")
    st.caption(
        f"混淆矩阵：TP {int(metrics.get('tp', 0))} / TN {int(metrics.get('tn', 0))} / "
        f"FP {int(metrics.get('fp', 0))} / FN {int(metrics.get('fn', 0))}"
        + (f" / Loss（损失值） {float(metrics.get('loss', 0.0)):.4f}" if "loss" in metrics else "")
    )


def render_model_assessment(result: dict) -> None:
    train_metrics = result.get("train_metrics", {}) if isinstance(result, dict) else {}
    val_metrics = result.get("best_val_metrics", {}) if isinstance(result, dict) else {}
    test_metrics = result.get("test_metrics", {}) if isinstance(result, dict) else {}
    messages = assessment_messages(train_metrics, val_metrics, test_metrics)
    if not messages:
        return
    st.markdown("##### 自动解读")
    for message in messages:
        st.write(f"- {message}")


def assessment_messages(train_metrics: dict, val_metrics: dict, test_metrics: dict) -> list[str]:
    messages: list[str] = []
    train_acc = safe_metric_value(train_metrics.get("accuracy"))
    val_acc = safe_metric_value(val_metrics.get("accuracy"))
    test_acc = safe_metric_value(test_metrics.get("accuracy"))
    if train_acc is None or val_acc is None or test_acc is None:
        return messages
    if train_acc - max(val_acc, test_acc) >= 0.1:
        messages.append("训练集明显高于验证集和测试集，存在过拟合风险。")
    if abs(val_acc - test_acc) <= 0.03:
        messages.append("验证集和测试集结果接近，当前模型的泛化稳定性较好。")
    elif abs(val_acc - test_acc) >= 0.1:
        messages.append("验证集和测试集差异较大，建议检查划分是否稳定，或增加样本后再评估。")
    if max(train_acc, val_acc, test_acc) < 0.7:
        messages.append("三组结果整体偏低，建议优先检查标签质量、文本字段选择和样本量。")
    if not messages:
        messages.append("当前三组结果没有明显异常，可以继续结合误判样本做业务复核。")
    return messages


def render_full_prediction_panel(
    dataset: Dataset,
    selected_dir: Path,
    asset_stats: dict[str, int],
    training_meta: dict,
) -> None:
    st.divider()
    st.markdown("#### 全量判定")
    st.caption("这里不是验证，而是用当前模型对原始导入数据 `raw_records` 做整表预测，输出完整判定结果。")
    if not asset_stats.get("raw", 0):
        st.info("当前没有原始数据，无法执行全量判定。")
        return
    job = load_full_prediction_job(dataset.dataset_id)
    source_columns = columns_for_table(dataset.dataset_id, "raw_records")
    default_text_column = str(training_meta.get("text_column", "") or (source_columns[0] if source_columns else ""))
    full_col1, full_col2 = st.columns(2)
    with full_col1:
        text_column = st.selectbox(
            "原始数据文本字段",
            source_columns,
            index=option_index(source_columns, default_text_column, 0) if source_columns else 0,
            key=f"full_prediction_text_{dataset.dataset_id}_{selected_dir.name}",
        ) if source_columns else ""
    with full_col2:
        batch_size = st.number_input(
            "判定批大小",
            min_value=1,
            max_value=256,
            value=int(training_meta.get("batch_size", 16) or 16),
            step=1,
            key=f"full_prediction_batch_{dataset.dataset_id}_{selected_dir.name}",
        )
    st.caption("全量判定不会改动模型参数，只会读取 `raw_records`，生成 `predictions` 结果表。")
    action_col1, action_col2 = st.columns(2)
    with action_col1:
        start_clicked = st.button(
            "开始全量判定",
            key=f"run_full_prediction_{dataset.dataset_id}_{selected_dir.name}",
            type="secondary",
            disabled=bool(job and job.get("status") == "running"),
        )
    with action_col2:
        stop_clicked = st.button(
            "停止全量判定",
            key=f"stop_full_prediction_{dataset.dataset_id}_{selected_dir.name}",
            disabled=not bool(job and job.get("status") == "running"),
        )

    if start_clicked:
        if not text_column:
            st.error("请先选择文本字段。")
            return
        existing_predictions = int(asset_stats.get("predictions", 0) or 0)
        if existing_predictions > 0:
            st.session_state[full_prediction_confirm_key(dataset.dataset_id)] = {
                "model_dir": str(selected_dir),
                "text_column": text_column,
                "batch_size": int(batch_size),
                "max_length": int(training_meta.get("max_length", 256) or 256),
                "total_rows": int(asset_stats.get("raw", 0)),
                "existing_predictions": existing_predictions,
            }
            st.rerun()
        start_full_prediction_job(
            dataset.dataset_id,
            {
                "model_dir": str(selected_dir),
                "text_column": text_column,
                "batch_size": int(batch_size),
                "max_length": int(training_meta.get("max_length", 256) or 256),
                "total_rows": int(asset_stats.get("raw", 0)),
            },
            resume=False,
        )
        set_full_prediction_autorun(dataset.dataset_id, True)
        st.rerun()

    if stop_clicked and job and job.get("status") == "running":
        set_full_prediction_autorun(dataset.dataset_id, False)
        job["status"] = "stopping"
        job["updated_at"] = datetime.now().isoformat()
        save_full_prediction_job(dataset.dataset_id, job)
        st.rerun()

    render_full_prediction_confirmation(dataset.dataset_id, asset_stats)

    job = load_full_prediction_job(dataset.dataset_id)
    if job:
        render_full_prediction_status(job)
        if job.get("status") == "stopping":
            job["status"] = "stopped"
            job["updated_at"] = datetime.now().isoformat()
            save_full_prediction_job(dataset.dataset_id, job)
            set_full_prediction_autorun(dataset.dataset_id, False)
            st.warning("全量判定已停止，已完成的结果保留在 predictions 表中。")
            return
        if job.get("status") == "running" and not is_full_prediction_autorun(dataset.dataset_id):
            st.info("检测到一个未完成的全量判定任务。当前不会自动继续，点击“继续全量判定”后才会恢复。")
            resume_col1, resume_col2 = st.columns(2)
            with resume_col1:
                continue_clicked = st.button("继续全量判定", key=f"continue_full_prediction_{dataset.dataset_id}", type="primary")
            with resume_col2:
                finish_stop_clicked = st.button("结束未完成任务", key=f"finish_full_prediction_{dataset.dataset_id}")
            if continue_clicked:
                set_full_prediction_autorun(dataset.dataset_id, True)
                st.rerun()
            if finish_stop_clicked:
                job["status"] = "stopped"
                job["updated_at"] = datetime.now().isoformat()
                save_full_prediction_job(dataset.dataset_id, job)
                st.rerun()
            render_partial_prediction_preview(dataset.dataset_id)
            return
        if job.get("status") == "running":
            try:
                updated_job = process_full_prediction_batch(dataset.dataset_id, job)
            except Exception as exc:
                job["status"] = "failed"
                job["error"] = str(exc)
                job["updated_at"] = datetime.now().isoformat()
                save_full_prediction_job(dataset.dataset_id, job)
                set_full_prediction_autorun(dataset.dataset_id, False)
                st.error(str(exc))
                return
            save_full_prediction_job(dataset.dataset_id, updated_job)
            if updated_job.get("status") == "completed":
                set_full_prediction_autorun(dataset.dataset_id, False)
                st.success(f"全量判定完成：已对原始数据 {int(updated_job.get('processed_rows', 0)):,} 行生成预测结果。")
                render_partial_prediction_preview(dataset.dataset_id)
                return
            st.rerun()
    render_partial_prediction_preview(dataset.dataset_id)


def full_prediction_job_path(dataset_id: str) -> Path:
    return STORE.dataset_dir(dataset_id) / "jobs" / "full_prediction_job.json"


def full_prediction_confirm_key(dataset_id: str) -> str:
    return f"full_prediction_confirm_{dataset_id}"


def full_prediction_autorun_key(dataset_id: str) -> str:
    return f"full_prediction_autorun_{dataset_id}"


def load_full_prediction_job(dataset_id: str) -> dict | None:
    path = full_prediction_job_path(dataset_id)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def save_full_prediction_job(dataset_id: str, payload: dict) -> None:
    path = full_prediction_job_path(dataset_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def clear_full_prediction_confirmation(dataset_id: str) -> None:
    st.session_state.pop(full_prediction_confirm_key(dataset_id), None)


def set_full_prediction_autorun(dataset_id: str, enabled: bool) -> None:
    st.session_state[full_prediction_autorun_key(dataset_id)] = bool(enabled)


def is_full_prediction_autorun(dataset_id: str) -> bool:
    return bool(st.session_state.get(full_prediction_autorun_key(dataset_id), False))


def render_full_prediction_confirmation(dataset_id: str, asset_stats: dict[str, int]) -> None:
    pending = st.session_state.get(full_prediction_confirm_key(dataset_id))
    if not pending:
        return
    st.warning(
        f"当前已有 {int(pending.get('existing_predictions', asset_stats.get('predictions', 0) or 0)):,} 行 predictions。"
        "开始新判定前，请确认是覆盖旧结果还是从现有进度续跑。"
    )
    action_col1, action_col2, action_col3 = st.columns(3)
    with action_col1:
        overwrite_clicked = st.button("覆盖旧 predictions", key=f"confirm_overwrite_{dataset_id}", type="primary")
    with action_col2:
        resume_clicked = st.button("续跑", key=f"confirm_resume_{dataset_id}")
    with action_col3:
        cancel_clicked = st.button("取消", key=f"confirm_cancel_{dataset_id}")

    if overwrite_clicked:
        start_full_prediction_job(dataset_id, pending, resume=False)
        clear_full_prediction_confirmation(dataset_id)
        set_full_prediction_autorun(dataset_id, True)
        st.rerun()
    if resume_clicked:
        start_full_prediction_job(dataset_id, pending, resume=True)
        clear_full_prediction_confirmation(dataset_id)
        set_full_prediction_autorun(dataset_id, True)
        st.rerun()
    if cancel_clicked:
        clear_full_prediction_confirmation(dataset_id)
        st.rerun()


def start_full_prediction_job(dataset_id: str, config: dict, resume: bool) -> None:
    total_rows = int(config.get("total_rows", 0) or 0)
    existing_job = load_full_prediction_job(dataset_id) or {}
    existing_predictions = int(get_asset_stats(STORE, dataset_id).get("predictions", 0) or 0)
    if resume:
        processed_rows = int(existing_job.get("processed_rows", existing_predictions) or 0)
        created_at = str(existing_job.get("created_at", datetime.now().isoformat()) or datetime.now().isoformat())
    else:
        clear_downstream_assets(STORE, dataset_id, "splits")
        processed_rows = 0
        created_at = datetime.now().isoformat()
    save_full_prediction_job(
        dataset_id,
        {
            "status": "running",
            "model_dir": str(config.get("model_dir", "")),
            "text_column": str(config.get("text_column", "")),
            "batch_size": int(config.get("batch_size", 16) or 16),
            "max_length": int(config.get("max_length", 256) or 256),
            "processed_rows": min(processed_rows, total_rows),
            "total_rows": total_rows,
            "last_batch_rows": 0,
            "created_at": created_at,
            "updated_at": datetime.now().isoformat(),
            "error": "",
        },
    )


def render_full_prediction_status(job: dict) -> None:
    processed = int(job.get("processed_rows", 0) or 0)
    total = int(job.get("total_rows", 0) or 0)
    percent = 0 if total <= 0 else min(100, int(processed / total * 100))
    status = str(job.get("status", "idle") or "idle")
    status_label = {
        "running": "执行中",
        "stopping": "停止中",
        "stopped": "已停止",
        "completed": "已完成",
        "failed": "失败",
    }.get(status, status)
    st.progress(percent / 100 if total > 0 else 0.0, text=f"全量判定进度 {percent}%")
    col1, col2, col3 = st.columns(3)
    col1.metric("当前状态", status_label)
    col2.metric("已处理行数", f"{processed:,}")
    col3.metric("总行数", f"{total:,}")
    if job.get("last_batch_rows"):
        st.caption(f"最近一批处理 {int(job.get('last_batch_rows', 0)):,} 行。")
    if job.get("error"):
        st.error(str(job.get("error")))


def process_full_prediction_batch(dataset_id: str, job: dict) -> dict:
    processed = int(job.get("processed_rows", 0) or 0)
    total = int(job.get("total_rows", 0) or 0)
    batch_size = int(job.get("batch_size", 16) or 16)
    model_dir = Path(str(job.get("model_dir", "")))
    text_column = str(job.get("text_column", "") or "")
    if processed >= total:
        job["status"] = "completed"
        job["updated_at"] = datetime.now().isoformat()
        return job
    records = fetch_table_records(STORE, dataset_id, "raw_records", limit=batch_size, offset=processed)
    if not records:
        job["status"] = "completed"
        job["updated_at"] = datetime.now().isoformat()
        return job
    prediction_rows = predict_records(
        records,
        PredictConfig(
            model_dir=model_dir,
            text_column=text_column,
            max_length=int(job.get("max_length", 256) or 256),
            batch_size=batch_size,
        ),
    )
    if processed == 0:
        replace_records_table(STORE, dataset_id, "predictions", prediction_rows)
    else:
        append_records_table(STORE, dataset_id, "predictions", prediction_rows)
    job["processed_rows"] = processed + len(prediction_rows)
    job["last_batch_rows"] = len(prediction_rows)
    job["updated_at"] = datetime.now().isoformat()
    if int(job["processed_rows"]) >= total:
        job["status"] = "completed"
    return job


def render_partial_prediction_preview(dataset_id: str) -> None:
    current_stats = get_asset_stats(STORE, dataset_id)
    prediction_count = int(current_stats.get("predictions", 0) or 0)
    if prediction_count <= 0:
        return
    st.markdown("##### 当前已完成结果预览")
    prediction_distribution = label_distribution(STORE, dataset_id, "predictions", "预测标签")
    positive_count = int(prediction_distribution.get("正样本", 0) or 0)
    negative_count = int(prediction_distribution.get("负样本", 0) or 0)
    other_count = max(0, prediction_count - positive_count - negative_count)
    stat_col1, stat_col2, stat_col3 = st.columns(3)
    stat_col1.metric("正样本", f"{positive_count:,}")
    stat_col2.metric("负样本", f"{negative_count:,}")
    stat_col3.metric("其他/空值", f"{other_count:,}")
    preview_limit = min(100, prediction_count)
    preview_rows = preview_table(STORE, dataset_id, "predictions", limit=preview_limit)
    st.caption(f"当前已写入 predictions：{prediction_count:,} 行，预览前 {preview_limit:,} 行。")
    st.dataframe(preview_rows, width="stretch", height=320)


def render_evaluation_details(title: str, rows: list[dict], key: str) -> None:
    st.markdown(f"##### {title}")
    if not rows:
        st.info("当前没有逐条评估结果。")
        return
    view_mode = st.radio("查看范围", ["全部样本", "仅误判样本"], horizontal=True, key=f"{key}_mode", label_visibility="collapsed")
    filtered = rows if view_mode == "全部样本" else [row for row in rows if not row.get("is_correct")]
    st.caption(f"{title}：共 {len(filtered):,} 条，默认展示前 200 条。")
    st.dataframe(filtered[:200], width="stretch", height=360)


def format_ratio(value: object) -> str:
    try:
        return f"{float(value):.4f}"
    except (TypeError, ValueError):
        return "-"


def metric_support(metrics: dict) -> int:
    try:
        support = int(metrics.get("support", 0))
    except (TypeError, ValueError):
        support = 0
    if support > 0:
        return support
    total = 0
    for key in ["tp", "tn", "fp", "fn"]:
        try:
            total += int(metrics.get(key, 0))
        except (TypeError, ValueError):
            continue
    return total


def safe_metric_value(value: object) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def render_step_outputs(dataset: Dataset, step: str) -> None:
    st.divider()
    st.markdown("#### 本步骤数据结果")
    toggle_key = f"show_outputs_{dataset.dataset_id}_{step}"
    if st.button("查看/收起本步骤产物", key=f"toggle_outputs_{dataset.dataset_id}_{step}"):
        st.session_state[toggle_key] = not st.session_state.get(toggle_key, False)
        st.rerun()
    if not st.session_state.get(toggle_key, False):
        return

    stats = get_asset_stats(STORE, dataset.dataset_id)
    if step == "model":
        result = latest_training_result(dataset.dataset_id)
        if result:
            st.json(result)
        else:
            st.info("还没有模型训练结果。")
        return
    if step == "predictions":
        result = latest_training_result(dataset.dataset_id)
        if result:
            st.json(result)
        if stats.get("predictions", 0) <= 0:
            if not result:
                st.info("还没有结果验证或全量判定产物。")
            return
        output_tables = [("全量判定结果", "predictions", stats.get("predictions", 0))]
        options = [f"{label} · {count:,} 行" for label, _, count in output_tables]
        selected = st.selectbox("选择要查看的产物", options, key=f"output_table_{dataset.dataset_id}_{step}")
        index = options.index(selected)
        label, table_name, count = output_tables[index]
        render_result_table(dataset, step, label, table_name, count)
        return

    output_tables = {
        "raw": [("原始数据", "raw_records", stats.get("raw", 0))],
        "sampled": [("样本池", "sampled_records", stats.get("sampled", 0))],
        "labeled": [
            ("标注集", "labeled_records", stats.get("labeled", 0)),
            ("高质量标注集", "reviewed_records", stats.get("reviewed", 0)),
        ],
        "splits": [
            ("训练集", "train_records", stats.get("train", 0)),
            ("验证集", "val_records", stats.get("val", 0)),
            ("测试集", "test_records", stats.get("test", 0)),
        ],
    }.get(step, [])

    available = [item for item in output_tables if item[2] > 0]
    if not available:
        st.info("当前步骤还没有可展示的数据产物。")
        return

    options = [f"{label} · {count:,} 行" for label, _, count in available]
    selected = st.selectbox("选择要查看的产物", options, key=f"output_table_{dataset.dataset_id}_{step}")
    index = options.index(selected)
    label, table_name, count = available[index]
    render_result_table(dataset, step, label, table_name, count)


def render_result_table(dataset: Dataset, step: str, label: str, table_name: str, row_count: int) -> None:
    page_size_key = f"output_page_size_{dataset.dataset_id}_{step}_{table_name}"
    page_size = st.selectbox("每页行数", [50, 100, 200, 500, 1000], index=1, key=page_size_key)
    max_page = max(1, (row_count + int(page_size) - 1) // int(page_size))
    page_key = f"output_page_{dataset.dataset_id}_{step}_{table_name}"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    st.session_state[page_key] = min(max(1, int(st.session_state[page_key])), max_page)
    render_pagination(page_key, max_page)

    page = int(st.session_state[page_key])
    offset = (page - 1) * int(page_size)
    preview = preview_table(STORE, dataset.dataset_id, table_name, limit=int(page_size), offset=offset)
    start = offset + 1 if row_count else 0
    end = min(offset + int(page_size), row_count)
    st.caption(f"{label}：第 {page} / {max_page} 页，显示 {start:,}-{end:,} / {row_count:,} 行。内部数据来自 DuckDB，不会一次性加载全量数据。")
    st.dataframe(preview, width="stretch", height=520)
    render_table_download(dataset, table_name, label, row_count)


def render_table_download(dataset: Dataset, table_name: str, label: str, row_count: int) -> None:
    export_key = f"excel_export_{dataset.dataset_id}_{table_name}_{row_count}"
    file_name = safe_file_name(f"{dataset.name}_{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    if st.button("生成 Excel 下载文件", key=f"prepare_download_{dataset.dataset_id}_{table_name}"):
        with st.spinner(f"正在导出 {row_count:,} 行到 Excel..."):
            try:
                st.session_state[export_key] = {
                    "file_name": file_name,
                    "data": export_table_to_excel_bytes(dataset.dataset_id, table_name),
                }
            except Exception as exc:
                st.error(str(exc))
                return
        st.success("Excel 文件已生成，可以下载。")
    export_payload = st.session_state.get(export_key)
    if export_payload:
        st.download_button(
            "下载完整 Excel",
            data=export_payload["data"],
            file_name=export_payload["file_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{dataset.dataset_id}_{table_name}_{row_count}",
        )


def export_table_to_excel_bytes(dataset_id: str, table_name: str, batch_size: int = 5000) -> bytes:
    import duckdb
    from openpyxl import Workbook

    db_path = STORE.duckdb_path(dataset_id)
    if not db_path.exists():
        raise ValueError("数据文件不存在，无法导出。")

    buffer = BytesIO()
    workbook = Workbook(write_only=True)
    max_excel_rows = 1_048_576
    sheet_index = 1
    rows_in_sheet = 0
    worksheet = None
    columns: list[str] = []

    def create_sheet():
        nonlocal sheet_index, rows_in_sheet, worksheet
        title = safe_sheet_title(table_name, sheet_index)
        worksheet = workbook.create_sheet(title=title)
        worksheet.append(columns)
        rows_in_sheet = 1
        sheet_index += 1

    with duckdb.connect(str(db_path)) as conn:
        cursor = conn.execute(f"SELECT * FROM {quote_ident(table_name)}")
        columns = [description[0] for description in cursor.description]
        create_sheet()
        while True:
            rows = cursor.fetchmany(batch_size)
            if not rows:
                break
            for row in rows:
                if rows_in_sheet >= max_excel_rows:
                    create_sheet()
                worksheet.append(["" if value is None else value for value in row])
                rows_in_sheet += 1

    workbook.save(buffer)
    return buffer.getvalue()


def safe_sheet_title(table_name: str, index: int) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in table_name)
    suffix = f"_{index}" if index > 1 else ""
    return (cleaned[: 31 - len(suffix)] + suffix) or f"sheet{index}"


def safe_file_name(file_name: str) -> str:
    invalid = set('/\\:*?"<>|')
    return "".join("_" if char in invalid else char for char in file_name)


def parse_field_specs(text: str) -> list[dict[str, str]]:
    fields: list[dict[str, str]] = []
    allowed_types = {"boolean", "string", "number", "integer", "enum"}
    for line in text.splitlines():
        if not line.strip():
            continue
        parts = [part.strip() for part in line.split("|", 2)]
        if len(parts) != 3:
            continue
        name, field_type, requirement = parts
        if not name.replace("_", "").isalnum() or "." in name:
            continue
        if field_type not in allowed_types:
            continue
        fields.append({"name": name, "type": field_type, "requirement": requirement})
    return fields


def normalize_field_rows(rows) -> list[dict[str, object]]:
    if hasattr(rows, "to_dict"):
        rows = rows.to_dict("records")
    fields: list[dict[str, object]] = []
    seen = set()
    for row in rows:
        name = str(row.get("name", "") or "").strip()
        if not name:
            continue
        field_type = FIELD_TYPE_LABELS.get(str(row.get("type_label", "") or "").strip(), "string")
        requirement = str(row.get("requirement", "") or "").strip()
        example = str(row.get("example", "") or "").strip()
        enum_options = str(row.get("enum_options", "") or "").strip()
        if not name.replace("_", "").isalnum() or "." in name or name in seen:
            continue
        seen.add(name)
        if field_type == "enum" and enum_options:
            requirement = f"{requirement}。可选值：{enum_options}" if requirement else f"可选值：{enum_options}"
        fields.append(
            {
                "name": name,
                "type": field_type,
                "type_label": FIELD_TYPE_VALUES.get(field_type, "文本"),
                "requirement": requirement or "按字段说明输出",
                "example": example,
                "required": bool(row.get("required", True)),
                "enum_options": enum_options,
            }
        )
    return fields


def field_rows_for_editor(fields: list[dict] | None) -> list[dict]:
    source = fields if fields else DEFAULT_LLM_FIELDS
    rows = []
    for field in source:
        field_type = field.get("type") or FIELD_TYPE_LABELS.get(str(field.get("type_label", "文本")), "string")
        rows.append(
            {
                "name": field.get("name", ""),
                "type_label": field.get("type_label") or FIELD_TYPE_VALUES.get(field_type, "文本"),
                "requirement": field.get("requirement", ""),
                "example": field.get("example", ""),
                "required": bool(field.get("required", True)),
                "enum_options": field.get("enum_options", ""),
            }
        )
    return rows


def build_structured_prompt(task_prompt: str, fields: list[dict[str, str]]) -> str:
    field_lines = []
    example = {}
    for field in fields:
        required_text = "必填" if field.get("required", True) else "可空"
        field_lines.append(f"- {field['name']} ({field['type']}，{required_text}): {field['requirement']}")
        example[field["name"]] = example_value(field["type"], str(field.get("example", "") or ""))
    return (
        "你是文本分类与信息抽取助手。\n\n"
        f"判定任务：\n{task_prompt.strip()}\n\n"
        "输出要求：\n"
        "- 严格输出 JSON 对象，不要输出 Markdown，不要输出解释文字。\n"
        "- JSON 只能包含下列一级字段，不允许嵌套对象，不允许数组，不允许增加字段。\n"
        "- 字段类型必须符合定义。\n\n"
        "字段定义：\n"
        + "\n".join(field_lines)
        + "\n\n输出 JSON 示例：\n"
        + str(example).replace("'", '"')
    )


def example_value(field_type: str, raw_example: str = ""):
    if raw_example:
        if field_type == "boolean":
            return raw_example.lower() in {"true", "1", "yes", "是", "对", "属于"}
        if field_type == "number":
            try:
                return float(raw_example)
            except ValueError:
                return raw_example
        if field_type == "integer":
            try:
                return int(float(raw_example))
            except ValueError:
                return raw_example
        return raw_example
    if field_type == "boolean":
        return True
    if field_type == "number":
        return 0.86
    if field_type == "integer":
        return 1
    return "示例"


def label_one_record(record: dict, config: LlmJudgeConfig, field_specs: list[dict[str, str]]) -> dict:
    output = dict(record)
    try:
        text = str(record.get(config.text_column, "") or "").strip()
        result = call_chat_json(config, text)
        allowed = {field["name"] for field in field_specs}
        for field in field_specs:
            output[field["name"]] = result.get(field["name"], "")
        output["LLM_JSON"] = str({key: result.get(key) for key in allowed})
        output["LLM_ERROR"] = ""
    except Exception as exc:
        for field in field_specs:
            output.setdefault(field["name"], "")
        output["LLM_JSON"] = ""
        output["LLM_ERROR"] = str(exc)
    return output


def columns_for_table(dataset_id: str, table_name: str) -> list[str]:
    rows = preview_table(STORE, dataset_id, table_name, limit=1)
    if not rows:
        return []
    return list(rows[0].keys())


def render_clear_panel(dataset: Dataset, compact: bool = False) -> None:
    if not compact:
        st.markdown("#### 清空/重置")
    st.warning("清空操作不可撤销。建议优先使用“只清空工作流产物，保留原始数据”。")
    scope_map = {
        "只清空工作流产物，保留原始数据": "derived",
        "清空模型": "models",
        "清空工作流历史": "history",
        "清空原始数据和导入记录": "raw",
        "清空当前数据集全部内容": "all",
    }
    choice = st.radio("清空范围", list(scope_map.keys()))
    confirm = st.text_input("输入数据集名称确认", placeholder=dataset.name)
    if st.button("执行清空", type="primary", disabled=confirm != dataset.name, key=f"clear_{dataset.dataset_id}_{compact}"):
        STORE.clear_dataset(dataset.dataset_id, scope_map[choice])
        st.success("清空完成。")
        st.rerun()


def render_history_panel(dataset: Dataset, compact: bool = False) -> None:
    if not compact:
        st.markdown("#### 导入历史")
    batches = STORE.list_import_batches(dataset.dataset_id, limit=50)
    if not batches:
        st.info("暂无导入记录。")
        return
    rows = [
        {
            "批次": item.batch_id,
            "文件": item.file_name,
            "行数": item.row_count,
            "状态": "导入成功" if item.status == "imported" else item.status,
            "时间": item.created_at,
            "错误": item.error_message,
        }
        for item in batches
    ]
    st.dataframe(rows, width="stretch")


def inject_styles() -> None:
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 2.25rem;
        }
        .jl-dataset-header {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(180px, 280px);
            align-items: start;
            gap: 32px;
            margin: 18px 0 56px;
        }
        .jl-dataset-header h2 {
            color: #26293a;
            font-size: 30px;
            line-height: 1.2;
            margin: 0 0 18px;
            letter-spacing: 0;
        }
        .jl-dataset-header p {
            color: #8a91a3;
            font-size: 16px;
            margin: 0;
        }
        .jl-metrics {
            display: grid;
            grid-template-columns: 1fr;
            gap: 26px;
        }
        .jl-metric span {
            display: block;
            color: #25283a;
            font-size: 16px;
            font-weight: 650;
            margin-bottom: 8px;
        }
        .jl-metric strong {
            display: block;
            color: #2b2d3f;
            font-size: 38px;
            font-weight: 500;
            line-height: 1;
            letter-spacing: 0;
        }
        .jl-flow-arrow {
            color: #8a98ad;
            font-size: 28px;
            font-weight: 500;
            text-align: center;
            line-height: 84px;
        }
        div[data-testid="stButton"] > button {
            border-radius: 8px;
        }
        div[data-testid="stButton"] > button[kind="primary"],
        div[data-testid="stFormSubmitButton"] > button[kind="primary"] {
            background: #2563eb !important;
            border-color: #2563eb !important;
            color: #ffffff !important;
        }
        div[data-testid="stButton"] > button[kind="primary"]:hover,
        div[data-testid="stFormSubmitButton"] > button[kind="primary"]:hover {
            background: #1d4ed8 !important;
            border-color: #1d4ed8 !important;
            color: #ffffff !important;
        }
        div[data-testid="stButton"] > button[kind="primary"]:focus,
        div[data-testid="stFormSubmitButton"] > button[kind="primary"]:focus {
            box-shadow: 0 0 0 0.2rem rgba(37, 99, 235, 0.24) !important;
        }
        div[data-baseweb="tag"] {
            background-color: #eff6ff !important;
            border-color: #bfdbfe !important;
            color: #1d4ed8 !important;
        }
        div[data-baseweb="tag"] span,
        div[data-baseweb="tag"] svg {
            color: #1d4ed8 !important;
            fill: #1d4ed8 !important;
        }
        div[data-baseweb="select"] input:focus,
        div[data-baseweb="input"] input:focus,
        div[data-baseweb="textarea"] textarea:focus {
            border-color: #2563eb !important;
        }
        .st-key-flow_nav div[data-testid="stButton"] > button {
            min-height: 112px;
            white-space: pre-line;
            line-height: 1.35;
            padding: 12px 8px;
            background: #f3f4f6 !important;
            border-color: #d1d5db !important;
            color: #6b7280 !important;
            box-shadow: none !important;
        }
        div[class*="st-key-flow_node_"][class*="_ready"] div[data-testid="stButton"] > button {
            background: #ecfdf3 !important;
            border-color: #86efac !important;
            color: #166534 !important;
        }
        div[class*="st-key-flow_node_"][class*="_active"] div[data-testid="stButton"] > button {
            background: #eff6ff !important;
            border-color: #3b82f6 !important;
            color: #1d4ed8 !important;
            box-shadow: 0 0 0 1px rgba(59, 130, 246, 0.18) !important;
        }
        div[class*="st-key-flow_node_"][class*="_notready"] div[data-testid="stButton"] > button {
            background: #f3f4f6 !important;
            border-color: #d1d5db !important;
            color: #9ca3af !important;
        }
        div[data-testid="stSlider"] {
            padding-top: 8px;
        }
        @media (max-width: 1180px) {
            .jl-flow-arrow {
                transform: rotate(90deg);
                font-size: 24px;
                line-height: 24px;
            }
        }
        @media (max-width: 760px) {
            .jl-dataset-header {
                grid-template-columns: 1fr;
                gap: 22px;
                margin-bottom: 34px;
            }
            .jl-metrics {
                grid-template-columns: 1fr 1fr;
            }
            .jl-metric strong {
                font-size: 30px;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
