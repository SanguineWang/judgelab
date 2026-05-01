from __future__ import annotations

import hashlib
import shutil
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Sequence

from openpyxl import load_workbook

from judgelab.data_io import dedupe_headers, normalize_header, repair_read_only_dimensions
from judgelab.workspace import WorkspaceStore, quote_ident, utc_now


@dataclass(frozen=True)
class ExcelFileProfile:
    path: Path
    file_name: str
    headers: List[str]
    row_count: int
    header_hash: str
    status: str
    missing_columns: List[str]
    extra_columns: List[str]
    message: str


@dataclass(frozen=True)
class PreflightReport:
    reference_headers: List[str]
    files: List[ExcelFileProfile]
    can_import_strict: bool
    total_rows: int


@dataclass(frozen=True)
class ImportOptions:
    allow_reordered_columns: bool = True
    allow_extra_columns: bool = False
    allow_missing_columns: bool = False
    copy_source_files: bool = True
    chunk_size: int = 1000


@dataclass(frozen=True)
class ImportSummary:
    dataset_id: str
    imported_files: int
    imported_rows: int
    skipped_files: int
    schema: List[str]


def preflight_excel_files(paths: Sequence[str | Path], reference_headers: Sequence[str] | None = None) -> PreflightReport:
    file_paths = [Path(path) for path in paths]
    if not file_paths:
        return PreflightReport(reference_headers=[], files=[], can_import_strict=False, total_rows=0)

    profiles_base = [read_excel_profile(path) for path in file_paths]
    reference = list(reference_headers or profiles_base[0].headers)
    files = [classify_profile(profile, reference) for profile in profiles_base]
    can_import_strict = all(file.status in {"matched", "reordered"} for file in files)
    total_rows = sum(file.row_count for file in files)
    return PreflightReport(reference_headers=reference, files=files, can_import_strict=can_import_strict, total_rows=total_rows)


def import_excel_files(
    store: WorkspaceStore,
    dataset_id: str,
    paths: Sequence[str | Path],
    options: ImportOptions,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> ImportSummary:
    existing_schema = store.get_raw_schema(dataset_id)
    report = preflight_excel_files(paths, existing_schema or None)
    if not report.files:
        return ImportSummary(dataset_id=dataset_id, imported_files=0, imported_rows=0, skipped_files=0, schema=[])

    blocked = [
        item
        for item in report.files
        if item.status == "missing_columns" and not options.allow_missing_columns
        or item.status == "extra_columns" and not options.allow_extra_columns
        or item.status == "empty_header"
    ]
    if blocked:
        names = "，".join(item.file_name for item in blocked)
        raise ValueError(f"导入预检未通过，请先处理表头问题: {names}")

    schema = list(existing_schema or report.reference_headers)
    if options.allow_extra_columns:
        for item in report.files:
            for column in item.extra_columns:
                if column not in schema:
                    schema.append(column)

    import duckdb

    db_path = store.duckdb_path(dataset_id)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    imported_rows = 0
    imported_files = 0
    total_rows = max(1, report.total_rows)

    with duckdb.connect(str(db_path)) as conn:
        ensure_dataset_tables(conn, schema)
        sync_dataset_schema(conn, schema)

        next_record_id = int(conn.execute("SELECT COALESCE(MAX(__record_id), 0) + 1 FROM raw_records").fetchone()[0])
        for profile in report.files:
            batch_id = f"batch_{uuid.uuid4().hex[:10]}"
            source_path = copy_source_file(store, dataset_id, profile.path) if options.copy_source_files else profile.path
            rows_inserted = import_one_excel(
                conn,
                profile.path,
                schema,
                batch_id,
                profile.file_name,
                next_record_id,
                options.chunk_size,
                progress_callback=progress_callback,
                progress_done=imported_rows,
                progress_total=total_rows,
            )
            next_record_id += rows_inserted
            imported_rows += rows_inserted
            imported_files += 1
            conn.execute(
                """
                INSERT INTO import_batches (
                    batch_id, file_name, file_path, file_hash, row_count, header_hash, status, created_at, error_message
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    profile.file_name,
                    str(source_path),
                    file_hash(profile.path),
                    rows_inserted,
                    profile.header_hash,
                    "imported",
                    utc_now(),
                    "",
                ),
            )

    store.update_dataset_stats(dataset_id, row_count=store.count_raw_records(dataset_id), column_count=len(schema), status="imported")
    return ImportSummary(
        dataset_id=dataset_id,
        imported_files=imported_files,
        imported_rows=imported_rows,
        skipped_files=0,
        schema=schema,
    )


def read_excel_profile(path: Path) -> ExcelFileProfile:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    repair_read_only_dimensions(worksheet)
    headers = dedupe_headers([normalize_header(worksheet.cell(row=1, column=col).value) for col in range(1, worksheet.max_column + 1)])
    headers = [header for header in headers if header]
    row_count = max(0, worksheet.max_row - 1)
    header_hash = hash_headers(headers)
    return ExcelFileProfile(
        path=path,
        file_name=path.name,
        headers=headers,
        row_count=row_count,
        header_hash=header_hash,
        status="pending",
        missing_columns=[],
        extra_columns=[],
        message="",
    )


def classify_profile(profile: ExcelFileProfile, reference_headers: Sequence[str]) -> ExcelFileProfile:
    reference = list(reference_headers)
    if not profile.headers:
        status = "empty_header"
        message = "表头为空"
    elif profile.headers == reference:
        status = "matched"
        message = "表头完全一致"
    else:
        missing = [header for header in reference if header not in profile.headers]
        extra = [header for header in profile.headers if header not in reference]
        if missing:
            status = "missing_columns"
            message = "缺少字段：" + "，".join(missing)
        elif extra:
            status = "extra_columns"
            message = "多出字段：" + "，".join(extra)
        else:
            status = "reordered"
            message = "字段一致但顺序不同，导入时会自动重排"
    return ExcelFileProfile(
        path=profile.path,
        file_name=profile.file_name,
        headers=profile.headers,
        row_count=profile.row_count,
        header_hash=profile.header_hash,
        status=status,
        missing_columns=[header for header in reference if header not in profile.headers],
        extra_columns=[header for header in profile.headers if header not in reference],
        message=message,
    )


def ensure_dataset_tables(conn: Any, schema: Sequence[str]) -> None:
    data_columns = ", ".join(f"{quote_ident(column)} VARCHAR" for column in schema)
    comma = ", " if data_columns else ""
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS raw_records (
            __record_id BIGINT,
            __batch_id VARCHAR,
            __source_file VARCHAR{comma}{data_columns}
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS import_batches (
            batch_id VARCHAR,
            file_name VARCHAR,
            file_path VARCHAR,
            file_hash VARCHAR,
            row_count BIGINT,
            header_hash VARCHAR,
            status VARCHAR,
            created_at VARCHAR,
            error_message VARCHAR
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dataset_schema (
            column_name VARCHAR,
            column_index INTEGER,
            data_type_guess VARCHAR,
            created_at VARCHAR
        )
        """
    )


def sync_dataset_schema(conn: Any, schema: Sequence[str]) -> None:
    conn.execute("DELETE FROM dataset_schema")
    rows = [(column, index, "text", utc_now()) for index, column in enumerate(schema, start=1)]
    conn.executemany(
        "INSERT INTO dataset_schema (column_name, column_index, data_type_guess, created_at) VALUES (?, ?, ?, ?)",
        rows,
    )


def import_one_excel(
    conn: Any,
    path: Path,
    schema: Sequence[str],
    batch_id: str,
    source_file: str,
    start_record_id: int,
    chunk_size: int,
    progress_callback: Callable[[int, int, str], None] | None = None,
    progress_done: int = 0,
    progress_total: int = 1,
) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    repair_read_only_dimensions(worksheet)
    headers = dedupe_headers([normalize_header(cell.value) for cell in worksheet[1]])
    header_index = {header: index for index, header in enumerate(headers) if header}
    insert_columns = ["__record_id", "__batch_id", "__source_file"] + list(schema)
    placeholders = ", ".join(["?"] * len(insert_columns))
    column_sql = ", ".join(quote_ident(column) for column in insert_columns)
    sql = f"INSERT INTO raw_records ({column_sql}) VALUES ({placeholders})"

    buffer: List[List[Any]] = []
    inserted = 0
    record_id = start_record_id
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        values = [record_id, batch_id, source_file]
        for column in schema:
            index = header_index.get(column)
            values.append("" if index is None or index >= len(row) or row[index] is None else str(row[index]))
        buffer.append(values)
        inserted += 1
        record_id += 1
        if len(buffer) >= chunk_size:
            conn.executemany(sql, buffer)
            buffer.clear()
            if progress_callback:
                progress_callback(progress_done + inserted, progress_total, f"正在导入 {source_file}")
    if buffer:
        conn.executemany(sql, buffer)
    if progress_callback:
        progress_callback(progress_done + inserted, progress_total, f"{source_file} 导入完成")
    return inserted


def copy_source_file(store: WorkspaceStore, dataset_id: str, path: Path) -> Path:
    target = store.imports_dir(dataset_id) / path.name
    if target.exists():
        target = store.imports_dir(dataset_id) / f"{path.stem}_{uuid.uuid4().hex[:6]}{path.suffix}"
    shutil.copy2(path, target)
    return target


def hash_headers(headers: Iterable[str]) -> str:
    payload = "\n".join(headers).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def file_hash(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()
