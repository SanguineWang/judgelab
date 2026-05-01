from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook, load_workbook


@dataclass(frozen=True)
class TableData:
    headers: List[str]
    rows: List[Dict[str, Any]]


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_table(path: str | Path) -> TableData:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return read_excel(file_path)
    if suffix == ".csv":
        return read_csv(file_path)
    raise ValueError(f"不支持的文件类型: {file_path.suffix}")


def read_excel(path: Path) -> TableData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    repair_read_only_dimensions(worksheet)
    headers = [normalize_header(worksheet.cell(row=1, column=col).value) for col in range(1, worksheet.max_column + 1)]
    headers = dedupe_headers(headers)

    rows: List[Dict[str, Any]] = []
    for row_index in range(2, worksheet.max_row + 1):
        record: Dict[str, Any] = {}
        has_value = False
        for col, header in enumerate(headers, start=1):
            if not header:
                continue
            value = worksheet.cell(row=row_index, column=col).value
            if value not in (None, ""):
                has_value = True
            record[header] = value
        if has_value:
            rows.append(record)
    return TableData(headers=[header for header in headers if header], rows=rows)


def read_csv(path: Path) -> TableData:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = dedupe_headers([normalize_header(header) for header in (reader.fieldnames or [])])
        rows = [{header: row.get(header) for header in headers if header} for row in reader]
    return TableData(headers=[header for header in headers if header], rows=rows)


def write_table(path: str | Path, rows: Sequence[Dict[str, Any]], headers: Sequence[str] | None = None) -> Path:
    file_path = Path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    output_headers = list(headers or infer_headers(rows))
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        write_csv(file_path, rows, output_headers)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        write_excel(file_path, rows, output_headers)
    else:
        raise ValueError(f"不支持的输出文件类型: {file_path.suffix}")
    return file_path


def write_excel(path: Path, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    workbook.save(path)


def write_csv(path: Path, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(headers))
        writer.writeheader()
        writer.writerows(rows)


def infer_headers(rows: Sequence[Dict[str, Any]]) -> List[str]:
    headers: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                headers.append(key)
                seen.add(key)
    return headers


def dedupe_headers(headers: Iterable[str]) -> List[str]:
    result: List[str] = []
    counts: Dict[str, int] = {}
    for header in headers:
        if not header:
            result.append("")
            continue
        counts[header] = counts.get(header, 0) + 1
        result.append(header if counts[header] == 1 else f"{header}_{counts[header]}")
    return result


def preview_rows(rows: Sequence[Dict[str, Any]], limit: int = 20) -> List[Dict[str, Any]]:
    return list(rows[: max(0, limit)])


def repair_read_only_dimensions(worksheet: Any) -> None:
    """Handle xlsx files whose sheet dimension says A1 even though data exists."""
    try:
        worksheet.reset_dimensions()
        worksheet.calculate_dimension(force=True)
    except (AttributeError, TypeError, ValueError):
        return
